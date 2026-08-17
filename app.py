import os
import queue
import re
import sys
import threading
import tkinter as tk
import csv
from datetime import datetime
from tkinter import ttk, filedialog, messagebox

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

from dotenv import load_dotenv

# Carga las variables del .env
load_dotenv()

from src.engine import cargar_claves, procesar, resumir

APP_TITLE = "Descargador de XML del SRI"

SRI_TIMEOUT = int(os.getenv("SRI_TIMEOUT", "30"))
CARPETA_DESCARGAS_DEFAULT = os.getenv("CARPETA_DESCARGAS", "./data/descargados")


def obtener_ruta_recurso(ruta_relativa):
    """Obtiene la ruta absoluta de un recurso, compatible con PyInstaller.
    Si la app corre empaquetada como .exe, los recursos viven en la carpeta
    temporal sys._MEIPASS; si corre como script normal, usan la carpeta actual.
    """
    if hasattr(sys, "_MEIPASS"):
        return os.path.join(sys._MEIPASS, ruta_relativa)
    return os.path.join(os.path.abspath("."), ruta_relativa)


class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title(APP_TITLE)
        self.geometry("760x640")
        self.minsize(680, 560)

        self.ruta_archivo_txt = tk.StringVar()

        self.ruc_propio = tk.StringVar()
        self.carpeta_destino = tk.StringVar(value=os.path.abspath(CARPETA_DESCARGAS_DEFAULT))
        self.hilos = tk.IntVar(value=4)
        self.pausa = tk.DoubleVar(value=0.0)
        self.mostrar_avanzado = tk.BooleanVar(value=False)

        self.resultados = []
        self._stop_flag = threading.Event()
        self._cola = queue.Queue()
        self._procesando = False

        self._build_ui()
        self._actualizar_boton_principal()
        self.after(100, self._poll_queue)

    def _build_ui(self):
        style = ttk.Style(self)
        try:
            style.theme_use(style.theme_use())
        except Exception:
            pass
        style.configure("Titulo.TLabel", font=("TkDefaultFont", 11, "bold"))
        style.configure("Grande.TButton", font=("TkDefaultFont", 11, "bold"))
        style.configure("Ok.TLabel", foreground="#1a7f37")
        style.configure("Error.TLabel", foreground="#b3261e")

        contenedor = ttk.Frame(self, padding=16)
        contenedor.pack(fill="both", expand=True)

        ttk.Label(contenedor, text="1. Elige tu archivo TXT del SRI", style="Titulo.TLabel").pack(anchor="w")
        ttk.Label(contenedor, text="Facturas Emitidas o Recibidas, según el listado exportado.",
                  foreground="#666").pack(anchor="w", pady=(0, 6))

        self._file_row(contenedor, "Archivo TXT seleccionado", self.ruta_archivo_txt, "archivo")

        ttk.Separator(contenedor).pack(fill="x", pady=12)

        ttk.Label(contenedor, text="2. Confirma el RUC", style="Titulo.TLabel").pack(anchor="w")

        frm_ruc = ttk.Frame(contenedor)
        frm_ruc.pack(anchor="w", pady=(0, 4))
        ttk.Entry(frm_ruc, textvariable=self.ruc_propio, width=20, font=("TkDefaultFont", 11)).pack(side="left")
        self.ruc_propio.trace_add("write", lambda *_: self._actualizar_boton_principal())

        ttk.Separator(contenedor).pack(fill="x", pady=12)

        ttk.Label(contenedor, text="3. Descarga", style="Titulo.TLabel").pack(anchor="w", pady=(0, 6))
        self.btn_principal = ttk.Button(
            contenedor, text="Descargar XML del SRI", style="Grande.TButton",
            command=self._iniciar_proceso,
        )
        self.btn_principal.pack(fill="x", ipady=8)

        self.lbl_ayuda_boton = ttk.Label(contenedor, text="", style="Error.TLabel")
        self.lbl_ayuda_boton.pack(anchor="w", pady=(4, 0))

        self.progress = ttk.Progressbar(contenedor, mode="determinate")
        self.progress.pack(fill="x", pady=(10, 4))

        self.lbl_estado = ttk.Label(contenedor, text="")
        self.lbl_estado.pack(anchor="w")

        self.frm_resultado = ttk.Frame(contenedor)
        self.frm_resultado.pack(fill="both", expand=True, pady=(10, 0))

        self.lbl_resumen = ttk.Label(self.frm_resultado, text="", font=("TkDefaultFont", 11, "bold"))
        self.lbl_resumen.pack(anchor="w")

        self.txt_detalle = tk.Text(self.frm_resultado, height=8, wrap="word", state="disabled",
                                    relief="flat", background="#f5f5f5")
        self.txt_detalle.pack(fill="both", expand=True, pady=(6, 6))

        frm_botones_final = ttk.Frame(self.frm_resultado)
        frm_botones_final.pack(anchor="w")
        self.btn_abrir_carpeta = ttk.Button(frm_botones_final, text="Abrir carpeta de resultados",
                                             command=self._abrir_carpeta, state="disabled")
        self.btn_abrir_carpeta.pack(side="left", padx=(0, 8))
        self.btn_reporte = ttk.Button(frm_botones_final, text="Guardar reporte (.xlsx)",
                                       command=self._exportar_reporte, state="disabled")
        self.btn_reporte.pack(side="left")

        ttk.Separator(contenedor).pack(fill="x", pady=(14, 6))
        chk = ttk.Checkbutton(contenedor, text="Opciones avanzadas", variable=self.mostrar_avanzado,
                               command=self._toggle_avanzado)
        chk.pack(anchor="w")

        self.frm_avanzado = ttk.Frame(contenedor)
        self._build_avanzado(self.frm_avanzado)

    def _build_avanzado(self, parent):
        pad = {"padx": 4, "pady": 4}
        ttk.Label(parent, text="Carpeta destino:").grid(row=0, column=0, sticky="w", **pad)
        ttk.Entry(parent, textvariable=self.carpeta_destino, width=42).grid(row=0, column=1, sticky="we", **pad)
        ttk.Button(parent, text="Elegir...", command=self._elegir_carpeta_destino).grid(row=0, column=2, **pad)

        ttk.Label(parent, text="Consultas simultáneas:").grid(row=1, column=0, sticky="w", **pad)
        ttk.Spinbox(parent, from_=1, to=10, textvariable=self.hilos, width=5).grid(row=1, column=1, sticky="w", **pad)

        ttk.Label(parent, text="Pausa entre solicitudes (seg):").grid(row=2, column=0, sticky="w", **pad)
        ttk.Spinbox(parent, from_=0, to=5, increment=0.1, textvariable=self.pausa, width=5).grid(row=2, column=1, sticky="w", **pad)

        ttk.Label(parent, text="Baja las consultas simultáneas o sube la pausa si el SRI empieza a fallar seguido.",
                  foreground="#666", wraplength=520, justify="left").grid(row=3, column=0, columnspan=3, sticky="w", **pad)
        parent.grid_columnconfigure(1, weight=1)

    def _toggle_avanzado(self):
        if self.mostrar_avanzado.get():
            self.frm_avanzado.pack(fill="x", pady=(4, 0))
        else:
            self.frm_avanzado.pack_forget()

    def _file_row(self, parent, label, var, kind):
        fila = ttk.Frame(parent)
        fila.pack(fill="x", pady=3)
        ttk.Label(fila, text=label, width=26).pack(side="left")
        entry = ttk.Entry(fila, textvariable=var)
        entry.pack(side="left", fill="x", expand=True, padx=6)
        ttk.Button(fila, text="Elegir archivo...", command=lambda v=var: self._elegir_txt(v)).pack(side="left")
        var.trace_add("write", lambda *_: self._actualizar_boton_principal())

    def _elegir_txt(self, var):
        path = filedialog.askopenfilename(title="Selecciona el archivo TXT", filetypes=[("Texto", "*.txt"), ("Todos", "*.*")])
        if path:
            var.set(path)
            if not self.ruc_propio.get().strip():
                m = re.match(r"(\d{10,13})", os.path.basename(path))
                if m:
                    self.ruc_propio.set(m.group(1))

    def _elegir_carpeta_destino(self):
        path = filedialog.askdirectory(title="Selecciona la carpeta destino")
        if path:
            self.carpeta_destino.set(path)

    def _abrir_carpeta(self):
        carpeta = self.carpeta_destino.get()
        os.makedirs(carpeta, exist_ok=True)
        try:
            if os.name == "nt":
                os.startfile(carpeta)
            elif os.uname().sysname == "Darwin":
                os.system(f'open "{carpeta}"')
            else:
                os.system(f'xdg-open "{carpeta}"')
        except Exception as e:
            messagebox.showinfo(APP_TITLE, f"Carpeta: {carpeta}\n(No se pudo abrir automáticamente: {e})")

    def _log_detalle(self, msg):
        self.txt_detalle.configure(state="normal")
        self.txt_detalle.insert("end", msg + "\n")
        self.txt_detalle.see("end")
        self.txt_detalle.configure(state="disabled")

    def _limpiar_detalle(self):
        self.txt_detalle.configure(state="normal")
        self.txt_detalle.delete("1.0", "end")
        self.txt_detalle.configure(state="disabled")

    def _actualizar_boton_principal(self):
        if self._procesando:
            return
        tiene_archivo = bool(self.ruta_archivo_txt.get().strip())
        tiene_ruc = bool(self.ruc_propio.get().strip())
        if tiene_archivo and tiene_ruc:
            self.btn_principal.config(state="normal")
            self.lbl_ayuda_boton.config(text="")
        else:
            self.btn_principal.config(state="disabled")
            faltantes = []
            if not tiene_archivo:
                faltantes.append("selecciona un archivo TXT")
            if not tiene_ruc:
                faltantes.append("ingresa tu RUC")
            self.lbl_ayuda_boton.config(text="Falta: " + " y ".join(faltantes))

    def _iniciar_proceso(self):
        e = self.ruta_archivo_txt.get().strip() or None
        r = None
        ruc = self.ruc_propio.get().strip()

        try:
            filas, mensajes = cargar_claves(e, r)
        except Exception as ex:
            messagebox.showerror(APP_TITLE, f"No se pudo leer el archivo:\n{ex}")
            return

        if not filas:
            messagebox.showwarning(APP_TITLE, "No se encontraron claves de acceso en el archivo seleccionado.")
            return

        self._limpiar_detalle()
        for m in mensajes:
            self._log_detalle(m)

        self._procesando = True
        self.btn_principal.config(state="disabled", text="Descargando...")
        self.btn_abrir_carpeta.config(state="disabled")
        self.btn_reporte.config(state="disabled")
        self.lbl_resumen.config(text="")
        self.progress.configure(maximum=len(filas), value=0)
        self.lbl_estado.config(text=f"Descargando 0 de {len(filas)}...")

        carpeta = self.carpeta_destino.get().strip()
        hilos = max(1, int(self.hilos.get() or 1))
        pausa = float(self.pausa.get() or 0)

        def on_item_done(item, idx, total):
            self._cola.put(("item", item, idx, total))

        def worker():
            resultados = procesar(
                filas, carpeta, ruc, hilos=hilos, pausa=pausa,
                timeout=SRI_TIMEOUT,
                on_item_done=on_item_done,
            )
            self._cola.put(("done", resultados))

        threading.Thread(target=worker, daemon=True).start()

    def _texto_excel(self, valor: str) -> str:
        """Fuerza a Excel/LibreOffice a tratar un valor como TEXTO, no como número.

        Solo se usa en la exportación a CSV: cualquier cadena de solo dígitos
        (RUC, claves de acceso, combinaciones como '2390057019001_1791290151001')
        se muestra en notación científica o pierde ceros a la izquierda si no
        se envuelve así. En XLSX este problema se resuelve distinto (ver
        _exportar_xlsx), así que ahí NO se usa este método.
        """
        valor = "" if valor is None else str(valor).strip()
        return f'="{valor}"' if valor else ""

    def _filas_reporte(self):
        """Arma el encabezado y las filas de datos, compartido entre CSV y XLSX."""
        encabezado = [
            "Clave de Acceso",
            "Origen TXT",
            "Descargado",
            "Categoría",
            "Tipo Documento",
            "Carpeta Contraparte",
            "Fecha Emisión",
            "Fecha Generación Reporte",
            "Estado / Error",
        ]
        fecha_reporte = datetime.now().strftime("%d/%m/%Y %H:%M:%S")

        filas = []
        for r in self.resultados:
            filas.append([
                getattr(r, "clave_acceso", "") or "",
                getattr(r, "origen", "") or "",
                "SÍ" if getattr(r, "ok", False) else "NO",
                getattr(r, "categoria", "") or "",
                getattr(r, "tipo", "") or "",
                getattr(r, "carpeta_id", "") or "",
                getattr(r, "fecha_emision", "") or "",
                fecha_reporte,
                getattr(r, "error", "") or "",
            ])
        return encabezado, filas

    def _exportar_reporte(self):
        if not self.resultados:
            return

        path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel", "*.xlsx"), ("CSV", "*.csv")],
            initialfile="reporte_descarga_sri.xlsx",
        )
        if not path:
            return

        if path.lower().endswith(".csv"):
            self._exportar_csv(path)
        else:
            self._exportar_xlsx(path)

        messagebox.showinfo(APP_TITLE, f"Reporte guardado en:\n{path}")

    def _exportar_csv(self, path):
        encabezado, filas = self._filas_reporte()
        with open(path, "w", newline="", encoding="utf-8-sig") as f:
            w = csv.writer(f, delimiter=";", quoting=csv.QUOTE_MINIMAL)
            w.writerow(encabezado)
            for fila in filas:
                # En CSV se necesita el truco ="..." para que las columnas
                # numéricas largas (clave de acceso, carpeta contraparte) no
                # se muestren en notación científica al abrir en Excel.
                fila_texto = list(fila)
                fila_texto[0] = self._texto_excel(fila_texto[0])  # Clave de Acceso
                fila_texto[5] = self._texto_excel(fila_texto[5])  # Carpeta Contraparte
                w.writerow(fila_texto)

    def _exportar_xlsx(self, path):
        encabezado, filas = self._filas_reporte()

        wb = Workbook()
        ws = wb.active
        ws.title = "Reporte SRI"

        # --- Encabezado grande, en negrita, con fondo y texto centrado ---
        ws.append(encabezado)
        ws.row_dimensions[1].height = 32
        fuente_encabezado = Font(bold=True, size=13, color="FFFFFF")
        relleno_encabezado = PatternFill(start_color=	"404040", end_color="404040", fill_type="solid")
        alineacion_encabezado = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for celda in ws[1]:
            celda.font = fuente_encabezado
            celda.fill = relleno_encabezado
            celda.alignment = alineacion_encabezado

        # --- Filas de datos ---
        for fila in filas:
            ws.append(fila)
            # Las columnas con dígitos largos se guardan explícitamente como
            # texto para que Excel no las muestre en notación científica.
            fila_idx = ws.max_row
            ws.cell(row=fila_idx, column=1).number_format = "@"  # Clave de Acceso
            ws.cell(row=fila_idx, column=6).number_format = "@"  # Carpeta Contraparte

        # --- Autoajustar ancho de columnas según el contenido más largo ---
        for col_idx, encabezado_col in enumerate(encabezado, start=1):
            letra = get_column_letter(col_idx)
            largo_max = len(str(encabezado_col))
            for fila in filas:
                valor = fila[col_idx - 1]
                largo_max = max(largo_max, len(str(valor)))
            ws.column_dimensions[letra].width = min(largo_max + 3, 60)

        ws.freeze_panes = "A2"  # el encabezado queda fijo al hacer scroll
        wb.save(path)

    def _poll_queue(self):
        try:
            while True:
                msg = self._cola.get_nowait()
                if msg[0] == "item":
                    _, item, idx, total = msg
                    self.progress.configure(value=idx)
                    self.lbl_estado.config(text=f"Descargando {idx} de {total}...")
                    if not item.ok:
                        self._log_detalle(f"✗ {item.clave_acceso} ({item.origen}): {item.error}")
                elif msg[0] == "done":
                    _, resultados = msg
                    self.resultados = resultados
                    self._mostrar_resumen(resultados)
        except queue.Empty:
            pass
        self.after(100, self._poll_queue)

    def _mostrar_resumen(self, resultados):
        self._procesando = False
        resumen = resumir(resultados)

        self.btn_principal.config(state="normal", text="Descargar XML del SRI")
        self.btn_abrir_carpeta.config(state="normal")
        self.btn_reporte.config(state="normal")
        self._actualizar_boton_principal()

        self.lbl_estado.config(text="Descarga finalizada.")
        self.lbl_resumen.config(
            text=(f"✔ {resumen['total_descargadas']} descargadas   "
                  f"✗ {resumen['total_no_descargadas']} no descargadas   "
                  f"(de {resumen['total_enviadas']} en total)")
        )

        self._limpiar_detalle()
        if resumen["por_tipo"]:
            self._log_detalle("Descargados por tipo:")
            for clave, cantidad in sorted(resumen["por_tipo"].items()):
                self._log_detalle(f"  • {clave}: {cantidad}")
        if resumen["errores"]:
            self._log_detalle("\nNo se pudieron descargar:")
            for clave, origen, err in resumen["errores"]:
                self._log_detalle(f"  • {clave} ({origen}): {err}")

        messagebox.showinfo(
            APP_TITLE,
            f"Descarga finalizada.\n\n"
            f"Descargadas: {resumen['total_descargadas']}\n"
            f"No descargadas: {resumen['total_no_descargadas']}\n"
            f"Total: {resumen['total_enviadas']}"
        )


if __name__ == "__main__":
    app = App()
    app.mainloop()